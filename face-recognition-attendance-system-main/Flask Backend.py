from flask import Flask, request, jsonify, render_template, send_file, redirect, session, url_for
from deepface import DeepFace
from pymongo import MongoClient
from pymongo.errors import PyMongoError
from dotenv import load_dotenv
import csv
import cv2
import os
import time
import openpyxl
import numpy as np
import base64
import re
import json
import uuid
from datetime import datetime
from io import BytesIO, StringIO
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = "securetrack-auth-secret"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env.mongo"))
attendance_filename = os.path.join(BASE_DIR, "attendance.xlsx")
pending_attendance_filename = os.path.join(BASE_DIR, "attendance_pending.json")
employee_registry_filename = os.path.join(BASE_DIR, "employee_registry.json")
daily_tasks_filename = os.path.join(BASE_DIR, "daily_tasks.json")
reference_images_path = os.path.join(BASE_DIR, "Images")
visitor_workbook_filename = os.path.join(BASE_DIR, "visitors.xlsx")
visitor_pending_filename = os.path.join(BASE_DIR, "visitor_pending.json")
visitor_faces_path = os.path.join(BASE_DIR, "VisitorFaces")
attendance_dict = {}

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"
MONGO_URI = os.environ.get("MONGO_URI", "mongodb://localhost:27017/")
MONGO_DB_NAME = os.environ.get("MONGO_DB_NAME", "userpassword")
MONGO_USERS_COLLECTION = os.environ.get("MONGO_USERS_COLLECTION", "users")
MONGO_ATTENDANCE_COLLECTION = os.environ.get("MONGO_ATTENDANCE_COLLECTION", "attendance_records")
MONGO_TASKS_COLLECTION = os.environ.get("MONGO_TASKS_COLLECTION", "daily_tasks")
DEFAULT_EMPLOYEE_ACCOUNT = {
    "username": "vishal",
    "password": "1234",
    "full_name": "Vishal",
    "employee_id": "EMP001",
    "email": "vishal@securetrack.local",
    "role": "employee",
}

mongo_client = MongoClient(MONGO_URI, serverSelectionTimeoutMS=3000)
mongo_db = mongo_client[MONGO_DB_NAME]
users_collection = mongo_db[MONGO_USERS_COLLECTION]
attendance_collection = mongo_db[MONGO_ATTENDANCE_COLLECTION]
tasks_collection = mongo_db[MONGO_TASKS_COLLECTION]

ATTENDANCE_HEADERS = [
    "Name",
    "Employee ID",
    "Date",
    "Check In",
    "Check Out",
    "Total Hours",
    "Status",
    "In Status",
    "Out Status",
    "Notes",
    "Confidence",
]

POLICY = {
    "office_start": "09:30",
    "buffer_time": "09:35",
    "early_exit": "16:50",
    "final_out": "18:00",
    "half_day_minutes": 240,
    "rescan_cooldown_seconds": 60,
}

DEFAULT_FACE_DISTANCE_THRESHOLD = 0.35

VISITOR_DATA_HEADERS = [
    "Name",
    "Purpose of Visit",
    "Person to Meet",
    "Question Responses",
    "Face Image Path",
    "Feedback",
    "Date",
    "Time",
    "Check Out Date",
    "Check Out Time",
    "Status",
]

VISITOR_QUESTION_HEADERS = [
    "Question ID",
    "Question Text",
]

DEFAULT_VISITOR_QUESTIONS = [
    {"id": "purpose_of_visit", "text": "What is the purpose of your visit?"},
    {"id": "person_to_meet", "text": "Whom do you want to meet?"},
]

# Initialize Excel for attendance tracking
def initialize_excel(filename):
    workbook = None
    try:
        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Attendance"
            sheet.append(ATTENDANCE_HEADERS)
            workbook.save(filename)
            return True, None

        workbook = openpyxl.load_workbook(filename)
        if "Attendance" not in workbook.sheetnames:
            sheet = workbook.create_sheet("Attendance")
            sheet.append(ATTENDANCE_HEADERS)
            workbook.save(filename)
            return True, None

        sheet = workbook["Attendance"]
        current_headers = [sheet.cell(row=1, column=idx).value for idx in range(1, len(ATTENDANCE_HEADERS) + 1)]
        if current_headers != ATTENDANCE_HEADERS:
            existing_rows = list(sheet.iter_rows(min_row=2, values_only=True))
            backup_filename = os.path.join(BASE_DIR, "attendance_backup_before_policy.xlsx")
            if not os.path.exists(backup_filename):
                workbook.save(backup_filename)

            workbook.remove(sheet)
            new_sheet = workbook.create_sheet("Attendance", 0)
            new_sheet.append(ATTENDANCE_HEADERS)

            for row in existing_rows:
                if not row or not row[0]:
                    continue
                legacy_name = row[0]
                legacy_date = row[1] if len(row) > 1 else ""
                legacy_check_in = row[2] if len(row) > 2 else ""
                legacy_status = row[3] if len(row) > 3 else "Present"
                new_sheet.append([
                    legacy_name,
                    "",
                    legacy_date or "",
                    legacy_check_in or "",
                    "",
                    "",
                    legacy_status or "Present",
                    "Legacy",
                    "",
                    "Migrated from old attendance format",
                    "",
                ])
            workbook.save(filename)
        return True, None
    except PermissionError:
        return False, "attendance.xlsx is open in Excel. Close it to allow Excel sync."
    except Exception as e:
        return False, str(e)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def ensure_sheet_headers(workbook, sheet_name, headers):
    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        return sheet

    sheet = workbook[sheet_name]
    current_headers = [sheet.cell(row=1, column=idx).value for idx in range(1, len(headers) + 1)]
    if current_headers != headers:
        existing_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.remove(sheet)
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for row in existing_rows:
            if not row:
                continue
            normalized = list(row[:len(headers)])
            while len(normalized) < len(headers):
                normalized.append("")
            sheet.append(normalized)
    return sheet

def initialize_visitor_workbook(filename):
    workbook = None
    try:
        os.makedirs(visitor_faces_path, exist_ok=True)
        if not os.path.exists(filename):
            workbook = openpyxl.Workbook()
            visitors_sheet = workbook.active
            visitors_sheet.title = "Visitors Data"
            visitors_sheet.delete_rows(1, visitors_sheet.max_row)
            visitors_sheet.append(VISITOR_DATA_HEADERS)
            questions_sheet = workbook.create_sheet("Questions")
            questions_sheet.append(VISITOR_QUESTION_HEADERS)
        else:
            workbook = openpyxl.load_workbook(filename)
            visitors_sheet = ensure_sheet_headers(workbook, "Visitors Data", VISITOR_DATA_HEADERS)
            questions_sheet = ensure_sheet_headers(workbook, "Questions", VISITOR_QUESTION_HEADERS)

        has_questions = any(
            (questions_sheet.cell(row=row_idx, column=1).value or "").strip()
            for row_idx in range(2, questions_sheet.max_row + 1)
        )
        if not has_questions:
            for question in DEFAULT_VISITOR_QUESTIONS:
                questions_sheet.append([question["id"], question["text"]])

        workbook.save(filename)
        return True, None
    except PermissionError:
        return False, "visitors.xlsx is open in Excel. Close it to allow visitor sync."
    except Exception as e:
        return False, str(e)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def load_pending_visitors():
    if not os.path.exists(visitor_pending_filename):
        return {}
    try:
        with open(visitor_pending_filename, "r", encoding="utf-8") as file:
            data = json.load(file)
            return data if isinstance(data, dict) else {}
    except Exception:
        return {}

def save_pending_visitors(records):
    with open(visitor_pending_filename, "w", encoding="utf-8") as file:
        json.dump(records, file, ensure_ascii=True, indent=2)

def register_pending_visitor(token, payload):
    records = load_pending_visitors()
    records[token] = payload
    save_pending_visitors(records)

def pop_pending_visitor(token):
    records = load_pending_visitors()
    payload = records.pop(token, None)
    save_pending_visitors(records)
    return payload

def get_visitor_questions():
    ok, error = initialize_visitor_workbook(visitor_workbook_filename)
    if not ok:
        raise RuntimeError(error)

    workbook = None
    try:
        workbook = openpyxl.load_workbook(visitor_workbook_filename)
        sheet = workbook["Questions"]
        questions = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            question_id = str(row[0] or "").strip() if len(row) > 0 else ""
            question_text = str(row[1] or "").strip() if len(row) > 1 else ""
            if not question_id or not question_text:
                continue
            questions.append({"id": question_id, "text": question_text})
        return questions or DEFAULT_VISITOR_QUESTIONS
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def serialize_question_answers(questions, answers):
    serialized = []
    for question in questions:
        answer_value = str(answers.get(question["id"], "")).strip()
        if answer_value:
            serialized.append(f'{question["text"]}: {answer_value}')
    return " | ".join(serialized)

def extract_named_answer(questions, answers, keyword_groups):
    for question in questions:
        question_text = question["text"].strip().lower()
        if any(all(keyword in question_text for keyword in group) for group in keyword_groups):
            return str(answers.get(question["id"], "")).strip()
    return ""

def save_visitor_record(payload):
    ok, error = initialize_visitor_workbook(visitor_workbook_filename)
    if not ok:
        return False, error

    workbook = None
    try:
        workbook = openpyxl.load_workbook(visitor_workbook_filename)
        visitors_sheet = workbook["Visitors Data"]
        questions_sheet = workbook["Questions"]
        questions = []
        for row in questions_sheet.iter_rows(min_row=2, values_only=True):
            question_id = str(row[0] or "").strip() if len(row) > 0 else ""
            question_text = str(row[1] or "").strip() if len(row) > 1 else ""
            if question_id and question_text:
                questions.append({"id": question_id, "text": question_text})
        if not questions:
            questions = DEFAULT_VISITOR_QUESTIONS
        answers = payload.get("answers", {})
        feedback = payload.get("feedback", {})

        purpose = extract_named_answer(questions, answers, [["purpose"], ["visit"]])
        person_to_meet = extract_named_answer(questions, answers, [["meet"], ["whom"]])
        feedback_summary = f'Rating: {feedback.get("rating", "")} | Comments: {feedback.get("comments", "")}'.strip(" |")

        visitors_sheet.append([
            payload.get("name", "").strip(),
            purpose,
            person_to_meet,
            serialize_question_answers(questions, answers),
            payload.get("face_image_path", ""),
            feedback_summary,
            payload.get("date", ""),
            payload.get("time", ""),
            "",
            "",
            "Checked In",
        ])
        workbook.save(visitor_workbook_filename)
        return True, None
    except PermissionError:
        return False, "visitors.xlsx is open in Excel. Please close it and try again."
    except Exception as e:
        return False, str(e)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def load_employee_registry():
    if not os.path.exists(employee_registry_filename):
        return {}
    try:
        with open(employee_registry_filename, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return {}

def save_employee_registry(registry):
    with open(employee_registry_filename, "w", encoding="utf-8") as file:
        json.dump(registry, file, ensure_ascii=True, indent=2)

def is_mongo_available():
    try:
        mongo_client.admin.command("ping")
        return True
    except PyMongoError:
        return False

def ensure_users_collection_indexes():
    try:
        users_collection.create_index("email", unique=True)
        users_collection.create_index("username", unique=True)
    except PyMongoError:
        pass

def ensure_attendance_collection_indexes():
    try:
        attendance_collection.create_index([("employee_id", 1), ("date", 1)], unique=True)
        attendance_collection.create_index("user_id")
    except PyMongoError:
        pass

def ensure_tasks_collection_indexes():
    try:
        tasks_collection.create_index("task_id", unique=True)
        tasks_collection.create_index("assigned_emp_id")
        tasks_collection.create_index("user_id")
    except PyMongoError:
        pass

def ensure_mongo_indexes():
    ensure_users_collection_indexes()
    ensure_attendance_collection_indexes()
    ensure_tasks_collection_indexes()

def get_employee_account_by_employee_id(employee_id):
    employee_id_value = (employee_id or "").strip()
    if not employee_id_value:
        return None
    try:
        return users_collection.find_one({
            "employee_id": employee_id_value,
            "role": "employee",
        })
    except PyMongoError:
        return None

def get_employee_account_by_user_id(user_id):
    user_id_value = str(user_id or "").strip()
    if not user_id_value or not is_mongo_available():
        return None
    try:
        from bson import ObjectId
        if ObjectId.is_valid(user_id_value):
            return users_collection.find_one({"_id": ObjectId(user_id_value), "role": "employee"})
    except Exception:
        pass
    return None

def get_employee_account_by_identifier(identifier):
    lookup_value = (identifier or "").strip().lower()
    if not lookup_value:
        return None
    try:
        return users_collection.find_one({
            "$or": [
                {"email": lookup_value},
                {"username": lookup_value},
            ],
            "role": "employee",
        })
    except PyMongoError:
        return None

def get_selected_role(value):
    role = (value or "").strip().lower()
    return role if role in {"admin", "employee"} else "employee"

def get_current_user():
    username = session.get("auth_username", "")
    employee_id = session.get("auth_employee_id", "")
    stored_user_id = session.get("auth_user_id", "")
    account = None
    if username:
        account = get_employee_account_by_identifier(username)
    if not account and employee_id:
        account = get_employee_account_by_employee_id(employee_id)

    return {
        "role": session.get("auth_role", ""),
        "username": username,
        "full_name": session.get("auth_full_name", ""),
        "employee_id": employee_id,
        "user_id": stored_user_id or str((account or {}).get("_id", "")),
    }

def get_employee_profile_context(current_user=None):
    current_user = current_user or get_current_user()
    username = (current_user.get("username") or "").strip()
    full_name = (current_user.get("full_name") or "").strip()
    employee_id = (current_user.get("employee_id") or "").strip()

    account = get_employee_account_by_identifier(username) if username else None
    profile = get_employee_profile(name=full_name, emp_id=employee_id)

    return {
        "full_name": full_name or (account or {}).get("name", ""),
        "username": username,
        "employee_id": employee_id or profile.get("emp_id", ""),
        "email": (account or {}).get("email", ""),
        "department": ((account or {}).get("department") or profile.get("department") or "General").strip() or "General",
        "purpose": (profile.get("purpose") or (account or {}).get("purpose") or "").strip(),
        "role": current_user.get("role", "employee") or "employee",
    }

def login_required(expected_role):
    current_role = session.get("auth_role")
    if current_role != expected_role:
        return False
    return True

def seed_default_employee_profile():
    registry = load_employee_registry()
    default_key = f"{DEFAULT_EMPLOYEE_ACCOUNT['username']}_auth_profile"
    if default_key not in registry:
        registry[default_key] = {
            "name": DEFAULT_EMPLOYEE_ACCOUNT["full_name"],
            "emp_id": DEFAULT_EMPLOYEE_ACCOUNT["employee_id"],
            "department": "General",
        }
        save_employee_registry(registry)

def seed_default_employee_account():
    if not is_mongo_available():
        return

    ensure_users_collection_indexes()
    existing_user = get_employee_account_by_identifier(DEFAULT_EMPLOYEE_ACCOUNT["username"])
    if existing_user:
        return

    try:
        users_collection.insert_one({
            "name": DEFAULT_EMPLOYEE_ACCOUNT["full_name"],
            "email": DEFAULT_EMPLOYEE_ACCOUNT["email"].strip().lower(),
            "username": DEFAULT_EMPLOYEE_ACCOUNT["username"].strip().lower(),
            "employee_id": DEFAULT_EMPLOYEE_ACCOUNT["employee_id"],
            "password": generate_password_hash(DEFAULT_EMPLOYEE_ACCOUNT["password"]),
            "role": "employee",
            "created_at": datetime.utcnow(),
        })
    except PyMongoError:
        pass

def create_employee_account(username, password, full_name, employee_id, email, department="General", purpose=""):
    username_key = (username or "").strip().lower()
    email_value = (email or "").strip().lower()
    department_value = (department or "General").strip() or "General"
    purpose_value = (purpose or "").strip()

    if not is_mongo_available():
        return False, "MongoDB is not available. Start MongoDB and try again."
    if not email_value:
        return False, "Email is required."
    if not username_key:
        username_key = email_value.split("@")[0].strip().lower()
    if not username_key:
        return False, "Username is required."
    if username_key == ADMIN_USERNAME:
        return False, "This username is reserved."
    ensure_mongo_indexes()

    try:
        existing_user = users_collection.find_one({
            "$or": [
                {"username": username_key},
                {"email": email_value},
            ]
        })
        if existing_user:
            if (existing_user.get("username") or "").strip().lower() == username_key:
                return False, "Username already exists."
            return False, "Email already exists."

        users_collection.insert_one({
            "name": (full_name or "").strip(),
            "email": email_value,
            "username": username_key,
            "employee_id": (employee_id or "").strip(),
            "password": generate_password_hash(password),
            "department": department_value,
            "purpose": purpose_value,
            "role": "employee",
            "created_at": datetime.utcnow(),
        })
    except PyMongoError as exc:
        return False, f"MongoDB error: {str(exc)}"

    registry = load_employee_registry()
    registry[f"{username_key}_auth_profile"] = {
        "name": (full_name or username_key).strip(),
        "emp_id": (employee_id or "").strip(),
        "department": department_value,
        "purpose": purpose_value,
        "email": email_value,
        "username": username_key,
    }
    save_employee_registry(registry)
    return True, "Employee account created successfully."

def update_employee_password(username, current_password, new_password):
    if not is_mongo_available():
        return False, "MongoDB is not available. Start MongoDB and try again."

    username_key = (username or "").strip().lower()
    account = get_employee_account_by_identifier(username_key)
    if not account:
        return False, "Employee account not found."
    if not check_password_hash(account.get("password") or "", current_password):
        return False, "Current password is incorrect."

    try:
        users_collection.update_one(
            {"_id": account["_id"]},
            {"$set": {"password": generate_password_hash(new_password)}}
        )
    except PyMongoError as exc:
        return False, f"MongoDB error: {str(exc)}"

    return True, "Password updated successfully."

def get_registered_personnel():
    registry = load_employee_registry()
    personnel_map = {}
    for item in registry.values():
        name = (item.get("name") or "").strip()
        emp_id = (item.get("emp_id") or "").strip()
        if not name and not emp_id:
            continue
        key = emp_id or name.lower()
        if key not in personnel_map:
            personnel_map[key] = {
                "name": name or "Unknown",
                "emp_id": emp_id,
                "role": "Employee",
                "department": (item.get("department") or "General").strip() or "General",
            }
    return sorted(personnel_map.values(), key=lambda person: ((person.get("name") or "").lower(), person.get("emp_id") or ""))

def get_employee_profile(name="", emp_id=""):
    registry = load_employee_registry()
    for item in registry.values():
        item_name = (item.get("name") or "").strip()
        item_emp_id = (item.get("emp_id") or "").strip()
        if emp_id and item_emp_id == emp_id:
            return item
        if name and item_name.lower() == str(name).strip().lower():
            return item
    return {}

def build_default_daily_tasks():
    personnel = get_registered_personnel()
    today_stamp = datetime.now().strftime("%Y-%m-%d")
    sample_titles = [
        ("Visitor Gate Verification", "Main Gate Security Desk", "09:00", "10:00", "In Progress"),
        ("Attendance Device Health Check", "Biometric Scanner Bay", "10:30", "11:30", "Pending"),
        ("Server Room Access Audit", "Floor 2 Server Room", "12:00", "13:00", "Completed"),
    ]
    tasks = []
    for idx, entry in enumerate(sample_titles, start=1):
        person = personnel[idx - 1] if idx - 1 < len(personnel) else {"name": "", "emp_id": "", "role": "Employee"}
        title, location, start_time, end_time, status = entry
        tasks.append({
            "id": f"task_{idx}_{int(time.time())}",
            "title": title,
            "location": location,
            "assigned_name": person.get("name", ""),
            "assigned_emp_id": person.get("emp_id", ""),
            "role": person.get("role", "Employee"),
            "date": today_stamp,
            "start_time": start_time,
            "end_time": end_time,
            "status": status,
            "notes": "Auto-created starter task",
            "created_at": datetime.now().isoformat(),
            "updated_at": datetime.now().isoformat(),
        })
    return tasks

def load_daily_tasks():
    if not os.path.exists(daily_tasks_filename):
        tasks = build_default_daily_tasks()
        save_daily_tasks(tasks)
        return tasks
    try:
        with open(daily_tasks_filename, "r", encoding="utf-8") as file:
            data = json.load(file)
            if isinstance(data, list):
                return data
    except Exception:
        pass
    tasks = build_default_daily_tasks()
    save_daily_tasks(tasks)
    return tasks

def save_daily_tasks(tasks):
    with open(daily_tasks_filename, "w", encoding="utf-8") as file:
        json.dump(tasks, file, ensure_ascii=True, indent=2)

def normalize_task_status(value):
    status = (value or "").strip().lower()
    mapping = {
        "pending": "Pending",
        "in progress": "In Progress",
        "completed": "Completed",
        "delayed": "Delayed",
    }
    return mapping.get(status, "Pending")

def upsert_daily_task(payload):
    tasks = load_daily_tasks()
    personnel = get_registered_personnel()
    personnel_lookup = {person.get("emp_id", ""): person for person in personnel if person.get("emp_id")}

    title = (payload.get("title") or "").strip()
    location = (payload.get("location") or "").strip()
    assigned_emp_id = (payload.get("assigned_emp_id") or "").strip()
    custom_name = (payload.get("assigned_name") or "").strip()
    role = (payload.get("role") or "").strip() or "Employee"
    date_value = (payload.get("date") or datetime.now().strftime("%Y-%m-%d")).strip()
    start_time = (payload.get("start_time") or "").strip()
    end_time = (payload.get("end_time") or "").strip()
    notes = (payload.get("notes") or "").strip()
    status = normalize_task_status(payload.get("status"))

    if not title:
        return False, "Task title is required", None
    if not location:
        return False, "Task location is required", None
    if not start_time:
        return False, "Start time is required", None

    assigned_person = personnel_lookup.get(assigned_emp_id, {})
    assigned_name = assigned_person.get("name") or custom_name
    if assigned_emp_id and not assigned_name:
        return False, "Selected employee is invalid", None
    if assigned_person.get("role"):
        role = assigned_person.get("role")

    task_id = (payload.get("id") or "").strip()
    now_iso = datetime.now().isoformat()
    task_record = {
        "id": task_id or f"task_{int(time.time() * 1000)}",
        "title": title,
        "location": location,
        "assigned_name": assigned_name,
        "assigned_emp_id": assigned_emp_id,
        "role": role,
        "date": date_value,
        "start_time": start_time,
        "end_time": end_time,
        "status": status,
        "notes": notes,
        "updated_at": now_iso,
    }

    existing_index = next((idx for idx, item in enumerate(tasks) if item.get("id") == task_record["id"]), None)
    if existing_index is None:
        task_record["created_at"] = now_iso
        tasks.append(task_record)
    else:
        task_record["created_at"] = tasks[existing_index].get("created_at", now_iso)
        tasks[existing_index] = task_record

    save_daily_tasks(tasks)
    sync_task_to_mongo(task_record)
    return True, "Task saved successfully", task_record

def delete_daily_task_record(task_id):
    tasks = load_daily_tasks()
    updated_tasks = [task for task in tasks if task.get("id") != task_id]
    if len(updated_tasks) == len(tasks):
        return False
    save_daily_tasks(updated_tasks)
    delete_task_from_mongo(task_id)
    return True

def register_employee_images(image_paths, name, emp_id, department="", email="", purpose="", username=""):
    registry = load_employee_registry()
    for path in image_paths:
        registry[os.path.basename(path)] = {
            "name": name,
            "emp_id": emp_id,
            "department": (department or "General").strip() or "General",
            "purpose": (purpose or "").strip(),
            "email": (email or "").strip().lower(),
            "username": (username or "").strip().lower(),
        }
    save_employee_registry(registry)

def get_employee_tasks(current_user=None):
    current_user = current_user or get_current_user()
    user_id = str(current_user.get("user_id") or "").strip()
    employee_id = (current_user.get("employee_id") or "").strip().lower()
    full_name = (current_user.get("full_name") or "").strip().lower()
    if is_mongo_available():
        try:
            query = {"$or": []}
            if user_id:
                query["$or"].append({"user_id": user_id})
            if employee_id:
                query["$or"].append({"assigned_emp_id": current_user.get("employee_id", "")})
            if full_name:
                query["$or"].append({"assigned_name_normalized": full_name})
            if query["$or"]:
                tasks = []
                for task in tasks_collection.find(query, {"_id": 0}).sort([("date", -1), ("start_time", -1)]):
                    tasks.append({
                        "id": task.get("task_id", ""),
                        "title": task.get("title", ""),
                        "location": task.get("location", ""),
                        "assigned_name": task.get("assigned_name", ""),
                        "assigned_emp_id": task.get("assigned_emp_id", ""),
                        "role": task.get("role", "Employee"),
                        "date": task.get("date", ""),
                        "start_time": task.get("start_time", ""),
                        "end_time": task.get("end_time", ""),
                        "status": task.get("status", "Pending"),
                        "notes": task.get("notes", ""),
                        "created_at": task.get("created_at", ""),
                        "updated_at": task.get("updated_at", ""),
                    })
                if tasks:
                    return tasks
        except PyMongoError:
            pass

    tasks = []
    for task in load_daily_tasks():
        assigned_emp_id = str(task.get("assigned_emp_id") or "").strip().lower()
        assigned_name = str(task.get("assigned_name") or "").strip().lower()
        if employee_id and assigned_emp_id == employee_id:
            tasks.append(task)
            continue
        if not assigned_emp_id and full_name and assigned_name == full_name:
            tasks.append(task)

    tasks.sort(key=lambda item: f"{item.get('date', '')} {item.get('start_time', '')}", reverse=True)
    return tasks

def sync_task_to_mongo(task):
    if not is_mongo_available() or not task:
        return
    ensure_tasks_collection_indexes()
    account = get_employee_account_by_employee_id(task.get("assigned_emp_id"))
    try:
        tasks_collection.update_one(
            {"task_id": task.get("id")},
            {"$set": {
                "task_id": task.get("id", ""),
                "title": task.get("title", ""),
                "location": task.get("location", ""),
                "assigned_name": task.get("assigned_name", ""),
                "assigned_name_normalized": str(task.get("assigned_name", "")).strip().lower(),
                "assigned_emp_id": task.get("assigned_emp_id", ""),
                "role": task.get("role", "Employee"),
                "date": task.get("date", ""),
                "start_time": task.get("start_time", ""),
                "end_time": task.get("end_time", ""),
                "status": task.get("status", "Pending"),
                "notes": task.get("notes", ""),
                "user_id": str((account or {}).get("_id", "")),
                "created_at": task.get("created_at", ""),
                "updated_at": task.get("updated_at", ""),
            }},
            upsert=True,
        )
    except PyMongoError:
        pass

def delete_task_from_mongo(task_id):
    if not is_mongo_available() or not task_id:
        return
    try:
        tasks_collection.delete_one({"task_id": task_id})
    except PyMongoError:
        pass

def sync_attendance_record_to_mongo(record):
    if not is_mongo_available() or not record:
        return
    employee_id = (record.get("emp_id") or "").strip()
    date_value = (record.get("date") or "").strip()
    if not employee_id or not date_value:
        return
    ensure_attendance_collection_indexes()
    account = get_employee_account_by_employee_id(employee_id)
    try:
        attendance_collection.update_one(
            {"employee_id": employee_id, "date": date_value},
            {"$set": {
                "user_id": str((account or {}).get("_id", "")),
                "name": record.get("name", ""),
                "employee_id": employee_id,
                "department": record.get("department", ""),
                "date": date_value,
                "check_in": record.get("check_in", "") or record.get("time", ""),
                "check_out": record.get("check_out", ""),
                "total_hours": record.get("total_hours", ""),
                "status": record.get("status", ""),
                "in_status": record.get("in_status", ""),
                "out_status": record.get("out_status", ""),
                "notes": record.get("notes", ""),
                "confidence": record.get("confidence", ""),
                "updated_at": datetime.utcnow(),
            }},
            upsert=True,
        )
    except PyMongoError:
        pass

def extract_identity_details(identity_path):
    file_name = os.path.basename(identity_path)
    registry = load_employee_registry()
    if file_name in registry:
        return registry[file_name].get("name", ""), registry[file_name].get("emp_id", "")

    raw_name = os.path.splitext(file_name)[0].replace("_", " ")
    tokens = os.path.splitext(file_name)[0].split("_")
    emp_id = ""
    if len(tokens) >= 2 and tokens[-1].isdigit():
        if len(tokens) >= 3:
            emp_id = tokens[-2]
            raw_name = " ".join(tokens[:-2])
    return raw_name.strip(), emp_id.strip()

def parse_match_distance(match_row):
    try:
        distance_value = match_row.get("distance")
    except AttributeError:
        distance_value = None
    if distance_value in (None, ""):
        return None
    try:
        return float(distance_value)
    except (TypeError, ValueError):
        return None

def parse_match_threshold(match_row):
    try:
        threshold_value = match_row.get("threshold")
    except AttributeError:
        threshold_value = None
    if threshold_value in (None, ""):
        return DEFAULT_FACE_DISTANCE_THRESHOLD
    try:
        return float(threshold_value)
    except (TypeError, ValueError):
        return DEFAULT_FACE_DISTANCE_THRESHOLD

def build_match_confidence(distance_value, threshold_value):
    if distance_value is None:
        return None
    baseline = threshold_value if threshold_value and threshold_value > 0 else 1.0
    confidence = 1 - (distance_value / baseline)
    return round(max(0.0, min(1.0, confidence)), 4)

def parse_policy_time(value):
    return datetime.strptime(value, "%H:%M").time()

def combine_date_time(date_str, time_str):
    return datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M:%S")

def format_duration(minutes):
    hours = max(0, minutes) // 60
    remainder = max(0, minutes) % 60
    return f"{hours}h {remainder:02d}m"

def calculate_work_minutes(check_in_value, check_out_value, break_minutes=0):
    if not check_in_value or not check_out_value:
        return 0
    check_in_dt = datetime.strptime(check_in_value, "%H:%M:%S")
    check_out_dt = datetime.strptime(check_out_value, "%H:%M:%S")
    total_minutes = int((check_out_dt - check_in_dt).total_seconds() // 60)
    return max(0, total_minutes - max(0, int(break_minutes or 0)))

def determine_in_status(check_in_time):
    office_start = parse_policy_time(POLICY["office_start"])
    buffer_time = parse_policy_time(POLICY["buffer_time"])
    if check_in_time <= office_start:
        return "On Time"
    if check_in_time <= buffer_time:
        return "Buffer"
    return "Late"

def determine_out_status(check_out_time):
    early_exit = parse_policy_time(POLICY["early_exit"])
    final_out = parse_policy_time(POLICY["final_out"])
    if check_out_time < early_exit:
        return "Early Exit"
    if check_out_time >= final_out:
        return "Final Out"
    return "Checked Out"

def determine_final_status(check_in_value, check_out_value, in_status, short_leave=False, break_minutes=0):
    if not check_out_value:
        return in_status

    work_minutes = calculate_work_minutes(check_in_value, check_out_value, break_minutes=break_minutes)
    check_out_time = datetime.strptime(check_out_value, "%H:%M:%S").time()
    if work_minutes < POLICY["half_day_minutes"]:
        return "Half Day"
    if check_out_time < parse_policy_time(POLICY["early_exit"]):
        return "Early Exit"
    if short_leave:
        return "Short Leave"
    if in_status == "Late":
        return "Late"
    if in_status == "Buffer":
        return "Buffer"
    return "Present"

def append_note(existing_note, new_note):
    existing_note = (existing_note or "").strip()
    if not existing_note:
        return new_note
    if new_note in existing_note:
        return existing_note
    return f"{existing_note}; {new_note}"

def split_notes(existing_note):
    return [part.strip() for part in str(existing_note or "").split(";") if part.strip()]

def join_notes(parts):
    return "; ".join(part for part in parts if part)

def get_note_tag(existing_note, tag_name):
    prefix = f"{tag_name}="
    for part in split_notes(existing_note):
        if part.startswith(prefix):
            return part[len(prefix):]
    return ""

def set_note_tag(existing_note, tag_name, value):
    prefix = f"{tag_name}="
    parts = [part for part in split_notes(existing_note) if not part.startswith(prefix)]
    parts.append(f"{prefix}{value}")
    return join_notes(parts)

def remove_note_tag(existing_note, tag_name):
    prefix = f"{tag_name}="
    parts = [part for part in split_notes(existing_note) if not part.startswith(prefix)]
    return join_notes(parts)

def parse_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default

def find_open_attendance_row(sheet, name, emp_id, date_value):
    for row_idx in range(sheet.max_row, 1, -1):
        row_name = sheet.cell(row=row_idx, column=1).value
        row_emp_id = sheet.cell(row=row_idx, column=2).value or ""
        row_date = sheet.cell(row=row_idx, column=3).value
        if row_name == name and row_emp_id == emp_id and row_date == date_value:
            return row_idx
    return None

def upsert_attendance_event(sheet, event):
    name = event["name"]
    emp_id = event.get("emp_id", "")
    date_value = event["date"]
    time_value = event["time"]
    confidence = event.get("confidence")
    event_dt = combine_date_time(date_value, time_value)
    row_idx = find_open_attendance_row(sheet, name, emp_id, date_value)

    if row_idx is None:
        in_status = determine_in_status(event_dt.time())
        sheet.append([
            name,
            emp_id,
            date_value,
            time_value,
            "",
            "",
            in_status,
            in_status,
            "",
            "",
            confidence if confidence is not None else "",
        ])
        return {
            "action": "check_in",
            "status": in_status,
            "in_status": in_status,
            "out_status": "",
            "check_in": time_value,
            "check_out": "",
            "total_hours": "",
            "notes": "",
        }

    check_in_value = sheet.cell(row=row_idx, column=4).value or ""
    check_out_value = sheet.cell(row=row_idx, column=5).value or ""
    existing_note = sheet.cell(row=row_idx, column=10).value or ""
    previous_confidence = sheet.cell(row=row_idx, column=11).value
    in_status = sheet.cell(row=row_idx, column=8).value or determine_in_status(event_dt.time())

    if check_in_value:
        check_in_dt = combine_date_time(date_value, check_in_value)
        cooldown_seconds = int((event_dt - check_in_dt).total_seconds())
        if not check_out_value and cooldown_seconds < POLICY["rescan_cooldown_seconds"]:
            note = append_note(existing_note, "Duplicate scan ignored")
            sheet.cell(row=row_idx, column=10).value = note
            return {
                "action": "ignored",
                "status": sheet.cell(row=row_idx, column=7).value or in_status,
                "in_status": in_status,
                "out_status": sheet.cell(row=row_idx, column=9).value or "",
                "check_in": check_in_value,
                "check_out": check_out_value,
                "total_hours": sheet.cell(row=row_idx, column=6).value or "",
                "notes": note,
            }

    if check_out_value:
        previous_out_dt = combine_date_time(date_value, check_out_value)
        if event_dt <= previous_out_dt:
            return {
                "action": "ignored",
                "status": sheet.cell(row=row_idx, column=7).value or in_status,
                "in_status": in_status,
                "out_status": sheet.cell(row=row_idx, column=9).value or "",
                "check_in": check_in_value,
                "check_out": check_out_value,
                "total_hours": sheet.cell(row=row_idx, column=6).value or "",
                "notes": existing_note,
            }

        note = append_note(existing_note, f"Short Leave: out {check_out_value}, returned {time_value}")
        sheet.cell(row=row_idx, column=5).value = ""
        sheet.cell(row=row_idx, column=6).value = ""
        sheet.cell(row=row_idx, column=7).value = "Short Leave"
        sheet.cell(row=row_idx, column=9).value = ""
        sheet.cell(row=row_idx, column=10).value = note
        if confidence is not None:
            sheet.cell(row=row_idx, column=11).value = confidence
        return {
            "action": "return",
            "status": "Short Leave",
            "in_status": in_status,
            "out_status": "",
            "check_in": check_in_value,
            "check_out": "",
            "total_hours": "",
            "notes": note,
        }

    out_status = determine_out_status(event_dt.time())
    break_minutes = parse_int(get_note_tag(existing_note, "BreakMinutes"), 0)
    work_minutes = calculate_work_minutes(check_in_value, time_value, break_minutes=break_minutes)
    total_hours = format_duration(work_minutes)
    short_leave = "Short Leave:" in existing_note
    final_status = determine_final_status(check_in_value, time_value, in_status, short_leave=short_leave, break_minutes=break_minutes)

    sheet.cell(row=row_idx, column=5).value = time_value
    sheet.cell(row=row_idx, column=6).value = total_hours
    sheet.cell(row=row_idx, column=7).value = final_status
    sheet.cell(row=row_idx, column=9).value = out_status
    if confidence is not None:
        sheet.cell(row=row_idx, column=11).value = confidence
    elif previous_confidence is not None:
        sheet.cell(row=row_idx, column=11).value = previous_confidence

    return {
        "action": "check_out",
        "status": final_status,
        "in_status": in_status,
        "out_status": out_status,
        "check_in": check_in_value,
        "check_out": time_value,
        "total_hours": total_hours,
        "notes": existing_note,
    }

def process_attendance_action(sheet, event, action):
    name = event["name"]
    emp_id = event.get("emp_id", "")
    date_value = event["date"]
    time_value = event["time"]
    row_idx = find_open_attendance_row(sheet, name, emp_id, date_value)

    if row_idx is None:
        return False, {
            "action": "error",
            "status": "error",
            "message": "No active attendance found for today. Mark attendance first.",
        }

    check_in_value = sheet.cell(row=row_idx, column=4).value or ""
    check_out_value = sheet.cell(row=row_idx, column=5).value or ""
    total_hours_value = sheet.cell(row=row_idx, column=6).value or ""
    current_status = sheet.cell(row=row_idx, column=7).value or ""
    in_status = sheet.cell(row=row_idx, column=8).value or ""
    out_status = sheet.cell(row=row_idx, column=9).value or ""
    existing_note = sheet.cell(row=row_idx, column=10).value or ""
    previous_confidence = sheet.cell(row=row_idx, column=11).value

    break_start_value = get_note_tag(existing_note, "BreakStart")
    break_minutes = parse_int(get_note_tag(existing_note, "BreakMinutes"), 0)

    if action == "check_out":
        if check_out_value:
            return False, {
                "action": "error",
                "status": current_status or "error",
                "message": f"{name} is already checked out for today.",
                "check_in": check_in_value,
                "check_out": check_out_value,
                "total_hours": total_hours_value,
                "notes": existing_note,
                "break_active": bool(break_start_value),
            }

        if break_start_value:
            break_started_dt = combine_date_time(date_value, break_start_value)
            checkout_dt = combine_date_time(date_value, time_value)
            if checkout_dt > break_started_dt:
                break_minutes += int((checkout_dt - break_started_dt).total_seconds() // 60)
            existing_note = remove_note_tag(existing_note, "BreakStart")
            existing_note = append_note(existing_note, f"Break ended {time_value} on checkout")

        existing_note = set_note_tag(existing_note, "BreakMinutes", break_minutes)
        work_minutes = calculate_work_minutes(check_in_value, time_value, break_minutes=break_minutes)
        total_hours = format_duration(work_minutes)
        short_leave = "Short Leave:" in existing_note
        resolved_out_status = determine_out_status(datetime.strptime(time_value, "%H:%M:%S").time())
        final_status = determine_final_status(check_in_value, time_value, in_status, short_leave=short_leave, break_minutes=break_minutes)

        sheet.cell(row=row_idx, column=5).value = time_value
        sheet.cell(row=row_idx, column=6).value = total_hours
        sheet.cell(row=row_idx, column=7).value = final_status
        sheet.cell(row=row_idx, column=9).value = resolved_out_status
        sheet.cell(row=row_idx, column=10).value = existing_note
        if event.get("confidence") is not None:
            sheet.cell(row=row_idx, column=11).value = event.get("confidence")
        elif previous_confidence is not None:
            sheet.cell(row=row_idx, column=11).value = previous_confidence

        return True, {
            "action": "check_out",
            "status": final_status,
            "message": f"Checked out {name} successfully.",
            "check_in": check_in_value,
            "check_out": time_value,
            "total_hours": total_hours,
            "notes": existing_note,
            "break_active": False,
        }

    if action == "toggle_break":
        if check_out_value:
            return False, {
                "action": "error",
                "status": current_status or "error",
                "message": f"{name} has already checked out for today.",
                "check_in": check_in_value,
                "check_out": check_out_value,
                "total_hours": total_hours_value,
                "notes": existing_note,
                "break_active": False,
            }

        if break_start_value:
            break_started_dt = combine_date_time(date_value, break_start_value)
            resume_dt = combine_date_time(date_value, time_value)
            if resume_dt <= break_started_dt:
                return False, {
                    "action": "error",
                    "status": current_status or "error",
                    "message": "Resume time must be after break start time.",
                    "check_in": check_in_value,
                    "check_out": "",
                    "total_hours": total_hours_value,
                    "notes": existing_note,
                    "break_active": True,
                }

            break_minutes += int((resume_dt - break_started_dt).total_seconds() // 60)
            updated_note = remove_note_tag(existing_note, "BreakStart")
            updated_note = set_note_tag(updated_note, "BreakMinutes", break_minutes)
            updated_note = append_note(updated_note, f"Break resumed {time_value}")
            sheet.cell(row=row_idx, column=7).value = in_status or "Present"
            sheet.cell(row=row_idx, column=10).value = updated_note
            return True, {
                "action": "resume_work",
                "status": sheet.cell(row=row_idx, column=7).value,
                "message": f"{name} resumed work.",
                "check_in": check_in_value,
                "check_out": "",
                "total_hours": "",
                "notes": updated_note,
                "break_active": False,
            }

        updated_note = set_note_tag(existing_note, "BreakStart", time_value)
        updated_note = set_note_tag(updated_note, "BreakMinutes", break_minutes)
        updated_note = append_note(updated_note, f"Break started {time_value}")
        sheet.cell(row=row_idx, column=7).value = "On Break"
        sheet.cell(row=row_idx, column=10).value = updated_note
        sheet.cell(row=row_idx, column=9).value = out_status
        return True, {
            "action": "take_break",
            "status": "On Break",
            "message": f"{name} is now on break.",
            "check_in": check_in_value,
            "check_out": "",
            "total_hours": "",
            "notes": updated_note,
            "break_active": True,
        }

    return False, {
        "action": "error",
        "status": "error",
        "message": "Unsupported attendance action.",
    }

def load_pending_attendance():
    if not os.path.exists(pending_attendance_filename):
        return []
    try:
        with open(pending_attendance_filename, "r", encoding="utf-8") as file:
            return json.load(file)
    except Exception:
        return []

def save_pending_attendance(records):
    with open(pending_attendance_filename, "w", encoding="utf-8") as file:
        json.dump(records, file, ensure_ascii=True, indent=2)

def build_attendance_record_from_row(row):
    if not row or not row[0]:
        return None

    name = row[0] or ""
    emp_id = ""
    date = ""
    time_value = ""
    status = ""
    confidence = None

    if len(row) >= 6:
        emp_id = row[1] or ""
        date = row[2] or ""
        time_value = row[3] or ""
        status = row[6] if len(row) > 6 else row[4] or ""
        confidence = row[10] if len(row) > 10 else row[5]
    elif len(row) == 5:
        emp_id = row[1] or ""
        date = row[2] or ""
        time_value = row[3] or ""
        status = row[4] or ""
    elif len(row) >= 4:
        date = row[1] or ""
        time_value = row[2] or ""
        status = row[3] or ""

    profile = get_employee_profile(name=name, emp_id=emp_id)
    account = get_employee_account_by_employee_id(emp_id)
    department = (profile.get("department") or "General").strip() or "General"
    display_status = "Queued" if str(status or "").lower() == "queued" else "Successful"

    return {
        "user_id": str((account or {}).get("_id", "")),
        "name": name,
        "emp_id": emp_id,
        "department": department,
        "date": date,
        "time": time_value,
        "check_in": row[3] if len(row) > 3 else time_value,
        "check_out": row[4] if len(row) > 4 else "",
        "total_hours": row[5] if len(row) > 5 else "",
        "status": status,
        "display_status": display_status,
        "in_status": row[7] if len(row) > 7 else "",
        "out_status": row[8] if len(row) > 8 else "",
        "notes": row[9] if len(row) > 9 else "",
        "confidence": confidence,
    }

def build_pending_attendance_record(record):
    profile = get_employee_profile(name=record.get("name", ""), emp_id=record.get("emp_id", ""))
    account = get_employee_account_by_employee_id(record.get("emp_id", ""))
    return {
        "user_id": str((account or {}).get("_id", "")),
        "name": record.get("name", ""),
        "emp_id": record.get("emp_id", ""),
        "department": (profile.get("department") or "General").strip() or "General",
        "date": record.get("date", ""),
        "time": record.get("time", ""),
        "check_in": record.get("time", ""),
        "check_out": "",
        "total_hours": "",
        "status": "Queued",
        "display_status": "Queued",
        "in_status": determine_in_status(datetime.strptime(record.get("time", "00:00:00"), "%H:%M:%S").time()) if record.get("time") else "",
        "out_status": "",
        "notes": "Pending Excel sync",
        "confidence": record.get("confidence"),
        "pending": True,
    }

def list_attendance_records(selected_date=None):
    records = []
    workbook = None
    try:
        flush_pending_attendance()
        ok, error = initialize_excel(attendance_filename)
        if not ok:
            raise RuntimeError(error)

        workbook = openpyxl.load_workbook(attendance_filename)
        sheet = workbook["Attendance"]

        for row in list(sheet.iter_rows(values_only=True))[1:]:
            record = build_attendance_record_from_row(row)
            if not record:
                continue
            if selected_date and str(record.get("date") or "") != selected_date:
                continue
            records.append(record)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

    for pending in load_pending_attendance():
        record = build_pending_attendance_record(pending)
        if selected_date and str(record.get("date") or "") != selected_date:
            continue
        records.append(record)

    return records

def record_matches_employee(record, current_user):
    record_emp_id = str(record.get("emp_id") or "").strip().lower()
    record_name = str(record.get("name") or "").strip().lower()
    employee_id = str(current_user.get("employee_id") or "").strip().lower()
    full_name = str(current_user.get("full_name") or "").strip().lower()
    username = str(current_user.get("username") or "").strip().lower()

    if employee_id and record_emp_id:
        return record_emp_id == employee_id
    if employee_id and not record_emp_id and full_name:
        return record_name == full_name
    if full_name and record_name:
        return record_name == full_name
    return bool(username and record_name == username)

def get_employee_attendance_history_records(current_user=None):
    current_user = current_user or get_current_user()
    if is_mongo_available():
        try:
            employee_id = (current_user.get("employee_id") or "").strip()
            query = {}
            if employee_id:
                query["employee_id"] = employee_id
            else:
                query["name"] = current_user.get("full_name", "")
            records = []
            for record in attendance_collection.find(query, {"_id": 0}).sort([("date", -1), ("check_in", -1)]):
                records.append({
                    "name": record.get("name", ""),
                    "emp_id": record.get("employee_id", ""),
                    "department": record.get("department", "General"),
                    "date": record.get("date", ""),
                    "time": record.get("check_in", ""),
                    "check_in": record.get("check_in", ""),
                    "check_out": record.get("check_out", ""),
                    "total_hours": record.get("total_hours", ""),
                    "status": record.get("status", ""),
                    "display_status": "Successful",
                    "in_status": record.get("in_status", ""),
                    "out_status": record.get("out_status", ""),
                    "notes": record.get("notes", ""),
                    "confidence": record.get("confidence", ""),
                })
            if records:
                return records
        except PyMongoError:
            pass
    records = [record for record in list_attendance_records() if record_matches_employee(record, current_user)]
    records.sort(key=lambda item: f"{item.get('date', '')} {item.get('check_in') or item.get('time', '')}", reverse=True)
    return records

def make_employee_attendance_history_csv(records):
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Name",
        "Employee ID",
        "Department",
        "Date",
        "Check In",
        "Check Out",
        "Total Hours",
        "Status",
        "In Status",
        "Out Status",
        "Notes",
        "Confidence",
    ])

    for record in records:
        writer.writerow([
            record.get("name", ""),
            record.get("emp_id", ""),
            record.get("department", ""),
            record.get("date", ""),
            record.get("check_in", ""),
            record.get("check_out", ""),
            record.get("total_hours", ""),
            record.get("status", ""),
            record.get("in_status", ""),
            record.get("out_status", ""),
            record.get("notes", ""),
            record.get("confidence", ""),
        ])

    return output.getvalue().encode("utf-8-sig")

def append_pending_attendance(record):
    records = load_pending_attendance()
    records.append(record)
    save_pending_attendance(records)

def flush_pending_attendance():
    pending_records = load_pending_attendance()
    if not pending_records:
        return True, 0, None

    try:
        initialize_excel(attendance_filename)
        workbook = openpyxl.load_workbook(attendance_filename)
        sheet = workbook["Attendance"]
        for record in pending_records:
            upsert_attendance_event(sheet, record)
        workbook.save(attendance_filename)
        workbook.close()
        save_pending_attendance([])
        return True, len(pending_records), None
    except PermissionError:
        return False, 0, "attendance.xlsx is open in Excel"
    except Exception as e:
        return False, 0, str(e)

excel_ready, excel_init_error = initialize_excel(attendance_filename)
if not excel_ready:
    print(f"Warning: {excel_init_error}")

visitor_excel_ready, visitor_excel_init_error = initialize_visitor_workbook(visitor_workbook_filename)
if not visitor_excel_ready:
    print(f"Warning: {visitor_excel_init_error}")

ensure_users_collection_indexes()
seed_default_employee_account()
seed_default_employee_profile()

# Load reference images
if os.path.exists(reference_images_path):
    reference_images = {}
    for file in os.listdir(reference_images_path):
        if file.endswith(".jpg") or file.endswith(".png"):
            name = os.path.splitext(file)[0].replace("_", " ")
            reference_images[name] = os.path.join(reference_images_path, file)
else:
    print(f"Warning: {reference_images_path} not found. Create it with face images.")

def recognize_identity_from_image(img_path):
    supported_extensions = {".jpg", ".jpeg", ".png"}
    registered_face_files = []

    if os.path.isdir(reference_images_path):
        for file_name in os.listdir(reference_images_path):
            extension = os.path.splitext(file_name)[1].lower()
            if extension in supported_extensions:
                registered_face_files.append(file_name)

    if not registered_face_files:
        return None

    try:
        results = DeepFace.find(img_path=img_path, db_path=reference_images_path, enforce_detection=False)
    except ValueError as exc:
        if "No item found" in str(exc):
            return None
        raise
    except Exception as exc:
        message = str(exc)
        if "No item found" in message or "does not exist" in message:
            return None
        raise

    if len(results) <= 0 or results[0].empty:
        return None

    matches = results[0]
    if "distance" in matches.columns:
        matches = matches.sort_values(by="distance", ascending=True)

    match = matches.iloc[0]
    distance = parse_match_distance(match)
    threshold = parse_match_threshold(match)
    if distance is None or distance > threshold:
        return None

    name, emp_id = extract_identity_details(match["identity"])
    confidence = build_match_confidence(distance, threshold)
    return {
        "name": name,
        "emp_id": emp_id,
        "confidence": confidence,
        "distance": distance,
        "threshold": threshold,
    }

@app.route('/')
def role_selection():
    return render_template('role_selection.html')

@app.route('/admin/dashboard')
@app.route('/index')
def admin_dashboard():
    if not login_required("admin"):
        return redirect(url_for('login', role='admin'))
    return render_template('index.html')

@app.route('/employee')
def employee():
    if not login_required("employee"):
        return redirect(url_for('login', role='employee'))
    return render_template('employee.html', current_user=get_current_user())

@app.route('/employee_attendance_history')
def employee_attendance_history():
    if not login_required("employee"):
        return redirect(url_for('login', role='employee'))
    return render_template('user/attendance_history.html', current_user=get_current_user())

@app.route('/user_tasks')
def user_tasks():
    if not login_required("employee"):
        return redirect(url_for('login', role='employee'))
    current_user = get_current_user()
    return render_template(
        'user/user_tasks.html',
        current_user=current_user,
        tasks=get_employee_tasks(current_user=current_user),
    )

@app.route('/employee_profile')
def employee_profile():
    if not login_required("employee"):
        return redirect(url_for('login', role='employee'))
    current_user = get_current_user()
    return render_template(
        'user/employee_profile.html',
        current_user=current_user,
        employee_profile=get_employee_profile_context(current_user),
    )

@app.route('/login')
def login():
    selected_role = get_selected_role(request.args.get('role'))
    return render_template('login.html', selected_role=selected_role)

@app.route('/signup')
def signup():
    selected_role = get_selected_role(request.args.get('role'))
    if selected_role != "employee":
        return redirect(url_for('login', role='employee'))
    return render_template('signup.html', selected_role=selected_role)

@app.route('/auth/login', methods=['POST'])
def auth_login():
    payload = request.get_json(silent=True) or {}
    selected_role = get_selected_role(payload.get("role"))
    username = (payload.get("username") or "").strip().lower()
    password = (payload.get("password") or "").strip()

    if not username or not password:
        return jsonify({"status": "error", "message": "Username and password are required."}), 400

    session.clear()

    if selected_role == "admin":
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            session["auth_role"] = "admin"
            session["auth_username"] = ADMIN_USERNAME
            session["auth_full_name"] = "Administrator"
            return jsonify({"status": "success", "redirect": url_for("admin_dashboard")})
        return jsonify({"status": "error", "message": "Invalid admin credentials."}), 401

    if not is_mongo_available():
        return jsonify({"status": "error", "message": "MongoDB is not available. Start MongoDB and try again."}), 500

    account = get_employee_account_by_identifier(username)
    if not account or not check_password_hash(account.get("password") or "", password):
        return jsonify({"status": "error", "message": "Invalid employee credentials."}), 401

    session["auth_role"] = "employee"
    session["auth_username"] = account.get("username") or username
    session["auth_full_name"] = account.get("name") or username.title()
    session["auth_employee_id"] = account.get("employee_id") or ""
    session["auth_user_id"] = str(account.get("_id") or "")
    return jsonify({"status": "success", "redirect": url_for("employee")})

@app.route('/auth/signup', methods=['POST'])
def auth_signup():
    payload = request.get_json(silent=True) or {}
    selected_role = get_selected_role(payload.get("role"))
    if selected_role != "employee":
        return jsonify({"status": "error", "message": "Only employees can sign up."}), 403

    full_name = (payload.get("full_name") or "").strip()
    employee_id = (payload.get("employee_id") or "").strip()
    department = (payload.get("department") or "General").strip() or "General"
    email = (payload.get("email") or "").strip().lower()
    password = payload.get("password") or ""

    if not full_name or not employee_id or not department or not email or not password:
        return jsonify({"status": "error", "message": "All fields are required."}), 400
    if len(password) < 4:
        return jsonify({"status": "error", "message": "Password must be at least 4 characters."}), 400

    success, message = create_employee_account(
        username="",
        password=password,
        full_name=full_name,
        employee_id=employee_id,
        email=email,
        department=department,
    )
    if not success:
        return jsonify({"status": "error", "message": message}), 400

    return jsonify({
        "status": "success",
        "message": message,
        "redirect": url_for("login", role="employee"),
    })

@app.route('/auth/change-password', methods=['POST'])
def change_employee_password():
    if not login_required("employee"):
        return jsonify({"status": "error", "message": "Please log in as employee first."}), 401

    payload = request.get_json(silent=True) or {}
    current_password = payload.get("current_password") or ""
    new_password = payload.get("new_password") or ""

    if not current_password or not new_password:
        return jsonify({"status": "error", "message": "Both password fields are required."}), 400
    if len(new_password) < 4:
        return jsonify({"status": "error", "message": "New password must be at least 4 characters."}), 400

    success, message = update_employee_password(session.get("auth_username"), current_password, new_password)
    if not success:
        return jsonify({"status": "error", "message": message}), 400

    return jsonify({"status": "success", "message": message})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('role_selection'))

@app.route('/new_register')
def new_register():
    return render_template('new_register.html')

@app.route('/attendance_log')
def attendance_log():
    return render_template('attendance_log.html')

@app.route('/visitor')
def visitor_log():
    return render_template('visitor.html')

@app.route('/get_visitor_questions')
def get_visitor_questions_route():
    try:
        return jsonify({"status": "success", "questions": get_visitor_questions()})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e), "questions": []}), 500

@app.route('/daily_task')
def daily_task():
    return render_template('daily_task.html')

@app.route('/get_daily_tasks')
def get_daily_tasks():
    try:
        tasks = load_daily_tasks()
        personnel = get_registered_personnel()
        tasks = sorted(tasks, key=lambda item: item.get("updated_at", ""), reverse=True)
        summary = {
            "active": len(tasks),
            "in_progress": sum(1 for task in tasks if normalize_task_status(task.get("status")) == "In Progress"),
            "delayed": sum(1 for task in tasks if normalize_task_status(task.get("status")) == "Delayed"),
            "completed": sum(1 for task in tasks if normalize_task_status(task.get("status")) == "Completed"),
        }
        return jsonify({
            "status": "success",
            "tasks": tasks,
            "personnel": personnel,
            "summary": summary,
        })
    except Exception as e:
        print(f"Error reading daily tasks: {e}")
        return jsonify({"status": "error", "message": "Unable to load daily tasks", "tasks": [], "personnel": []})

@app.route('/save_daily_task', methods=['POST'])
def save_daily_task_route():
    payload = request.get_json(silent=True) or request.form.to_dict()
    success, message, task = upsert_daily_task(payload)
    if not success:
        return jsonify({"status": "error", "message": message}), 400
    return jsonify({"status": "success", "message": message, "task": task})

@app.route('/delete_daily_task', methods=['POST'])
def delete_daily_task_route():
    payload = request.get_json(silent=True) or request.form.to_dict()
    task_id = (payload.get("id") or "").strip()
    if not task_id:
        return jsonify({"status": "error", "message": "Task id is required"}), 400
    if not delete_daily_task_record(task_id):
        return jsonify({"status": "error", "message": "Task not found"}), 404
    return jsonify({"status": "success", "message": "Task deleted"})

@app.route('/export_attendance')
def export_attendance():
    return send_file(attendance_filename, as_attachment=True)

@app.route('/export_visitors')
def export_visitors():
    ok, error = initialize_visitor_workbook(visitor_workbook_filename)
    if not ok:
        return jsonify({"status": "error", "message": error}), 500
    return send_file(visitor_workbook_filename, as_attachment=True)

@app.route('/visitor_face')
def visitor_face():
    relative_path = (request.args.get('path') or '').strip().replace("/", os.sep)
    if not relative_path:
        return jsonify({"status": "error", "message": "Image path is required"}), 400

    absolute_path = os.path.abspath(os.path.join(BASE_DIR, relative_path))
    visitor_faces_root = os.path.abspath(visitor_faces_path)

    if not absolute_path.startswith(visitor_faces_root + os.sep) and absolute_path != visitor_faces_root:
        return jsonify({"status": "error", "message": "Invalid image path"}), 400

    if not os.path.exists(absolute_path):
        return jsonify({"status": "error", "message": "Image not found"}), 404

    return send_file(absolute_path)

@app.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    if 'image' not in request.files:
        return jsonify({"status": "error", "message": "No file part"})

    file = request.files['image']
    if file.filename == '':
        return jsonify({"status": "error", "message": "No selected file"})

    img_path = "temp_image.png"
    file.save(img_path)

    try:
        identity = recognize_identity_from_image(img_path)

        if identity:
            name = identity["name"]
            emp_id = identity["emp_id"]
            confidence = identity.get("confidence")
            profile = get_employee_profile(name=name, emp_id=emp_id)
            department = (profile.get("department") or "General").strip() or "General"

            timestamp = time.strftime("%Y-%m-%d %H:%M:%S").split()
            record = {
                "name": name,
                "emp_id": emp_id,
                "date": timestamp[0],
                "time": timestamp[1],
                "confidence": confidence
            }

            flush_pending_attendance()
            save_success, save_error, attendance_result = save_attendance_event(attendance_filename, record)
            if not save_success:
                append_pending_attendance(record)
                attendance_result = {
                    "action": "queued",
                    "status": "Queued",
                    "in_status": "",
                    "out_status": "",
                    "check_in": record["time"],
                    "check_out": "",
                    "total_hours": "",
                    "notes": "Excel file is open",
                }

            with open(img_path, "rb") as img_file:
                img_base64 = base64.b64encode(img_file.read()).decode('utf-8')
            os.remove(img_path)

            response = {
                "status": "success",
                "name": name,
                "emp_id": emp_id,
                "department": department,
                "message": f"Attendance marked for {name}",
                "image": img_base64,
                "date": timestamp[0],
                "time": timestamp[1],
                "confidence": confidence,
                "attendance_action": attendance_result.get("action"),
                "attendance_label": attendance_result.get("status"),
                "check_in": attendance_result.get("check_in"),
                "check_out": attendance_result.get("check_out"),
                "total_hours": attendance_result.get("total_hours"),
                "notes": attendance_result.get("notes"),
            }
            if not save_success:
                response["message"] = f"Attendance marked for {name}. Excel file is open, so the entry is queued and will sync automatically."
                response["queued"] = True
                response["queue_error"] = save_error

            return jsonify(response)
        else:
            token = uuid.uuid4().hex
            timestamp = datetime.now()
            file_name = f"visitor_{timestamp.strftime('%Y%m%d_%H%M%S')}_{token}.png"
            saved_face_path = os.path.join(visitor_faces_path, file_name)
            os.makedirs(visitor_faces_path, exist_ok=True)

            with open(img_path, "rb") as img_file:
                image_bytes = img_file.read()
            img_base64 = base64.b64encode(image_bytes).decode('utf-8')

            with open(saved_face_path, "wb") as visitor_file:
                visitor_file.write(image_bytes)

            relative_face_path = os.path.relpath(saved_face_path, BASE_DIR).replace("\\", "/")
            register_pending_visitor(token, {
                "face_image_path": relative_face_path,
                "created_at": timestamp.isoformat(),
            })

            os.remove(img_path)
            return jsonify({
                "status": "visitor_required",
                "message": "Visitor detected. Please complete the visitor form and feedback before finishing.",
                "visitor_token": token,
                "visitor_image": img_base64,
                "visitor_image_path": relative_face_path,
                "questions": get_visitor_questions(),
                "date": timestamp.strftime("%Y-%m-%d"),
                "time": timestamp.strftime("%H:%M:%S"),
            })
    except Exception as e:
        if os.path.exists(img_path):
            os.remove(img_path)
        return jsonify({"status": "error", "message": f"Error: {str(e)}"})

@app.route('/attendance_action', methods=['POST'])
def attendance_action():
    action = (request.form.get('action') or '').strip().lower()
    if action not in ['check_out', 'toggle_break']:
        return jsonify({"status": "error", "message": "Invalid attendance action"})

    if 'image' not in request.files:
        return jsonify({"status": "error", "message": "No file part"})

    file = request.files['image']
    if file.filename == '':
        return jsonify({"status": "error", "message": "No selected file"})

    img_path = "temp_action_image.png"
    file.save(img_path)

    try:
        identity = recognize_identity_from_image(img_path)
        if not identity:
            return jsonify({"status": "error", "message": "Face not recognized"})

        timestamp = time.strftime("%Y-%m-%d %H:%M:%S").split()
        record = {
            "name": identity["name"],
            "emp_id": identity["emp_id"],
            "date": timestamp[0],
            "time": timestamp[1],
            "confidence": identity.get("confidence")
        }

        flush_pending_attendance()
        initialize_excel(attendance_filename)
        workbook = openpyxl.load_workbook(attendance_filename)
        sheet = workbook["Attendance"]
        success, result = process_attendance_action(sheet, record, action)
        workbook.save(attendance_filename)
        workbook.close()
        profile = get_employee_profile(name=record.get("name", ""), emp_id=record.get("emp_id", ""))
        sync_attendance_record_to_mongo({
            "name": record.get("name", ""),
            "emp_id": record.get("emp_id", ""),
            "department": (profile.get("department") or "General").strip() or "General",
            "date": record.get("date", ""),
            "time": record.get("time", ""),
            "check_in": result.get("check_in", ""),
            "check_out": result.get("check_out", ""),
            "total_hours": result.get("total_hours", ""),
            "status": result.get("status", ""),
            "in_status": result.get("in_status", ""),
            "out_status": result.get("out_status", ""),
            "notes": result.get("notes", ""),
            "confidence": record.get("confidence", ""),
        })

        response = {
            "status": "success" if success else "error",
            "name": record["name"],
            "emp_id": record["emp_id"],
            "message": result.get("message"),
            "attendance_action": result.get("action"),
            "attendance_label": result.get("status"),
            "check_in": result.get("check_in"),
            "check_out": result.get("check_out"),
            "total_hours": result.get("total_hours"),
            "notes": result.get("notes"),
            "break_active": result.get("break_active", False),
            "date": timestamp[0],
            "time": timestamp[1],
            "confidence": identity.get("confidence"),
        }
        return jsonify(response), (200 if success else 400)
    except PermissionError:
        return jsonify({"status": "error", "message": "attendance.xlsx is open in Excel. Close it and try again."}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500
    finally:
        if os.path.exists(img_path):
            os.remove(img_path)

@app.route('/submit_visitor', methods=['POST'])
def submit_visitor():
    try:
        token = (request.form.get('visitor_token') or '').strip()
        visitor_name = (request.form.get('visitor_name') or '').strip()
        answers_raw = request.form.get('answers', '{}')
        feedback_rating = (request.form.get('feedback_rating') or '').strip()
        feedback_comments = (request.form.get('feedback_comments') or '').strip()

        if not token:
            return jsonify({"status": "error", "message": "Visitor token is required"}), 400
        if not visitor_name:
            return jsonify({"status": "error", "message": "Visitor name is required"}), 400
        if not feedback_rating or not feedback_comments:
            return jsonify({"status": "error", "message": "Feedback rating and comments are required"}), 400

        try:
            answers = json.loads(answers_raw)
        except Exception:
            return jsonify({"status": "error", "message": "Invalid visitor answers"}), 400

        if not isinstance(answers, dict):
            return jsonify({"status": "error", "message": "Invalid visitor answers"}), 400

        questions = get_visitor_questions()
        missing_questions = [
            question["text"]
            for question in questions
            if not str(answers.get(question["id"], "")).strip()
        ]
        if missing_questions:
            return jsonify({"status": "error", "message": "All visitor questions are required"}), 400

        pending_visitor = pop_pending_visitor(token)
        if not pending_visitor:
            return jsonify({"status": "error", "message": "Visitor session expired. Please scan again."}), 400

        timestamp = datetime.now()
        success, error = save_visitor_record({
            "name": visitor_name,
            "answers": answers,
            "feedback": {
                "rating": feedback_rating,
                "comments": feedback_comments,
            },
            "face_image_path": pending_visitor.get("face_image_path", ""),
            "date": timestamp.strftime("%Y-%m-%d"),
            "time": timestamp.strftime("%H:%M:%S"),
        })
        if not success:
            register_pending_visitor(token, pending_visitor)
            return jsonify({"status": "error", "message": error}), 500

        return jsonify({
            "status": "success",
            "message": f"Visitor entry saved for {visitor_name}.",
            "face_image_path": pending_visitor.get("face_image_path", ""),
            "date": timestamp.strftime("%Y-%m-%d"),
            "time": timestamp.strftime("%H:%M:%S"),
        })
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500

@app.route('/register_face', methods=['POST'])
def register_face():
    name = request.form.get('name', '')
    emp_id = request.form.get('emp_id', '')
    department = request.form.get('department', 'General')
    email = (request.form.get('email') or '').strip().lower()
    password = request.form.get('password') or ''
    purpose = (request.form.get('purpose') or '').strip()
    username = (request.form.get('username') or '').strip().lower()
    if not name or not emp_id or not email or not password:
        return jsonify({"status": "error", "message": "Name, Employee ID, email, and password are required"})

    files = request.files.getlist('images')
    if not files:
        single_file = request.files.get('image')
        if single_file and single_file.filename:
            files = [single_file]

    valid_files = [file for file in files if file and file.filename]
    if len(valid_files) < 5:
        return jsonify({"status": "error", "message": "At least 5 face images are required"})

    if len(password) < 4:
        return jsonify({"status": "error", "message": "Password must be at least 4 characters"})
    if not username:
        username = (email.split("@")[0] if email else "").strip().lower()
        if not username:
            username = re.sub(r'[^A-Za-z0-9]+', '', emp_id).lower()

    success, message = create_employee_account(
        username=username,
        password=password,
        full_name=name,
        employee_id=emp_id,
        email=email,
        department=department,
        purpose=purpose,
    )
    if not success:
        return jsonify({"status": "error", "message": message})

    os.makedirs(reference_images_path, exist_ok=True)

    safe_name = re.sub(r'[^A-Za-z0-9]+', '_', name.strip()).strip('_') or "user"
    safe_emp_id = re.sub(r'[^A-Za-z0-9]+', '_', emp_id.strip()).strip('_') or "emp"
    prefix = f"{safe_name}_{safe_emp_id}_"

    existing_indices = []
    for file_name in os.listdir(reference_images_path):
        if file_name.startswith(prefix) and file_name.lower().endswith((".jpg", ".png", ".jpeg")):
            suffix = os.path.splitext(file_name)[0].replace(prefix, "")
            if suffix.isdigit():
                existing_indices.append(int(suffix))

    next_index = max(existing_indices, default=0) + 1
    saved_paths = []

    for offset, file in enumerate(valid_files):
        img_path = os.path.join(reference_images_path, f"{prefix}{next_index + offset}.jpg")
        file.save(img_path)
        saved_paths.append(img_path)

    register_employee_images(saved_paths, name, emp_id, department=department, email=email, purpose=purpose, username=username)
    print(f"Registered {name} (ID: {emp_id}) - saved {len(saved_paths)} images")
    return jsonify({
        "status": "success",
        "message": f"Face and login registered for {name} (ID: {emp_id}) with {len(saved_paths)} images",
        "emp_id": emp_id,
        "department": department,
        "email": email,
        "purpose": purpose,
        "saved_count": len(saved_paths)
    })

@app.route('/get_attendance')
def get_attendance():
    try:
        selected_date = (request.args.get('date') or datetime.now().strftime("%Y-%m-%d")).strip()
        data = list_attendance_records(selected_date=selected_date)
        return jsonify({"status": "success", "attendance": data[-50:], "selected_date": selected_date})
    except Exception as e:
        print(f"Error reading attendance: {e}")
        return jsonify({"status": "error", "message": "No attendance data", "attendance": []})

@app.route('/admin_manual_checkout', methods=['POST'])
def admin_manual_checkout():
    workbook = None
    if not login_required("admin"):
        return jsonify({"status": "error", "message": "Admin access required."}), 403

    try:
        payload = request.get_json(silent=True) or {}
        user_id = str(payload.get("user_id") or "").strip()
        selected_date = (payload.get("date") or datetime.now().strftime("%Y-%m-%d")).strip()

        if not user_id:
            return jsonify({"status": "error", "message": "User ID is required."}), 400

        account = get_employee_account_by_user_id(user_id)
        if not account:
            return jsonify({"status": "error", "message": "Employee account not found."}), 404

        employee_name = (account.get("name") or "").strip()
        employee_id = (account.get("employee_id") or "").strip()
        if not employee_name and not employee_id:
            return jsonify({"status": "error", "message": "Employee attendance identity is incomplete."}), 400

        flush_pending_attendance()
        ok, error = initialize_excel(attendance_filename)
        if not ok:
            return jsonify({"status": "error", "message": error}), 500

        workbook = openpyxl.load_workbook(attendance_filename)
        sheet = workbook["Attendance"]
        current_time = datetime.now().strftime("%H:%M:%S")

        success, result = process_attendance_action(sheet, {
            "name": employee_name,
            "emp_id": employee_id,
            "date": selected_date,
            "time": current_time,
        }, "check_out")

        if not success:
            return jsonify({"status": "error", "message": result.get("message", "Unable to check out user.")}), 400

        workbook.save(attendance_filename)
        profile = get_employee_profile(name=employee_name, emp_id=employee_id)
        sync_attendance_record_to_mongo({
            "name": employee_name,
            "emp_id": employee_id,
            "department": (profile.get("department") or "General").strip() or "General",
            "date": selected_date,
            "check_in": result.get("check_in", ""),
            "check_out": result.get("check_out", ""),
            "total_hours": result.get("total_hours", ""),
            "status": result.get("status", ""),
            "in_status": determine_in_status(datetime.strptime(result.get("check_in", "00:00:00"), "%H:%M:%S").time()) if result.get("check_in") else "",
            "out_status": determine_out_status(datetime.strptime(result.get("check_out", "00:00:00"), "%H:%M:%S").time()) if result.get("check_out") else "",
            "notes": result.get("notes", ""),
        })

        return jsonify({
            "status": "success",
            "message": f"{employee_name or employee_id} checked out successfully.",
            "user_id": user_id,
            "employee_id": employee_id,
            "date": selected_date,
            "check_out": result.get("check_out", ""),
            "total_hours": result.get("total_hours", ""),
            "attendance_status": result.get("status", ""),
        })
    except PermissionError:
        return jsonify({"status": "error", "message": "attendance.xlsx is open in Excel. Close it and try again."}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

@app.route('/get_employee_attendance_history')
def get_employee_attendance_history():
    if not login_required("employee"):
        return jsonify({"status": "error", "message": "Please log in as employee first.", "attendance": []}), 401

    try:
        current_user = get_current_user()
        records = get_employee_attendance_history_records(current_user=current_user)
        return jsonify({
            "status": "success",
            "attendance": records,
            "employee": {
                "full_name": current_user.get("full_name", ""),
                "employee_id": current_user.get("employee_id", ""),
            },
        })
    except Exception as e:
        print(f"Error reading employee attendance history: {e}")
        return jsonify({"status": "error", "message": "Unable to load attendance history", "attendance": []}), 500

@app.route('/get_employee_tasks')
def get_employee_tasks_route():
    if not login_required("employee"):
        return jsonify({"status": "error", "message": "Please log in as employee first.", "tasks": []}), 401

    try:
        current_user = get_current_user()
        return jsonify({"status": "success", "tasks": get_employee_tasks(current_user=current_user)})
    except Exception as e:
        print(f"Error reading employee tasks: {e}")
        return jsonify({"status": "error", "message": "Unable to load assigned tasks", "tasks": []}), 500

@app.route('/download_employee_attendance_csv')
def download_employee_attendance_csv():
    if not login_required("employee"):
        return redirect(url_for('login', role='employee'))

    try:
        current_user = get_current_user()
        records = get_employee_attendance_history_records(current_user=current_user)
        csv_bytes = make_employee_attendance_history_csv(records)
        safe_employee_id = re.sub(r'[^A-Za-z0-9_-]+', '_', current_user.get("employee_id") or current_user.get("username") or "employee")
        return send_file(
            BytesIO(csv_bytes),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'attendance_history_{safe_employee_id}.csv',
        )
    except Exception as e:
        print(f"Error exporting employee attendance CSV: {e}")
        return jsonify({"status": "error", "message": "Unable to export attendance CSV"}), 500

@app.route('/get_visitors')
def get_visitors():
    workbook = None
    try:
        ok, error = initialize_visitor_workbook(visitor_workbook_filename)
        if not ok:
            return jsonify({"status": "error", "message": error, "visitors": []}), 500

        workbook = openpyxl.load_workbook(visitor_workbook_filename)
        sheet = workbook["Visitors Data"]
        data = []

        for row_index, row in enumerate(list(sheet.iter_rows(values_only=True))[1:], start=2):
            if not row or not row[0]:
                continue

            data.append({
                "row_id": row_index,
                "name": row[0] or "",
                "purpose": row[1] or "",
                "person_to_meet": row[2] or "",
                "question_responses": row[3] or "",
                "face_image_path": row[4] or "",
                "feedback": row[5] or "",
                "date": row[6] or "",
                "time": row[7] or "",
                "check_out_date": row[8] if len(row) > 8 else "",
                "check_out_time": row[9] if len(row) > 9 else "",
                "status": row[10] if len(row) > 10 else "Checked In",
            })

        data.sort(key=lambda item: f'{item.get("date", "")} {item.get("time", "")}', reverse=True)
        return jsonify({"status": "success", "visitors": data})
    except Exception as e:
        print(f"Error reading visitors: {e}")
        return jsonify({"status": "error", "message": "Unable to load visitor data", "visitors": []}), 500
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

@app.route('/visitor_checkout', methods=['POST'])
def visitor_checkout():
    workbook = None
    try:
        row_id_raw = (request.form.get('row_id') or '').strip()
        if not row_id_raw.isdigit():
            return jsonify({"status": "error", "message": "Invalid visitor record"}), 400

        row_id = int(row_id_raw)
        ok, error = initialize_visitor_workbook(visitor_workbook_filename)
        if not ok:
            return jsonify({"status": "error", "message": error}), 500

        workbook = openpyxl.load_workbook(visitor_workbook_filename)
        sheet = workbook["Visitors Data"]

        if row_id < 2 or row_id > sheet.max_row:
            return jsonify({"status": "error", "message": "Visitor record not found"}), 404

        visitor_name = str(sheet.cell(row=row_id, column=1).value or "").strip()
        if not visitor_name:
            return jsonify({"status": "error", "message": "Visitor record not found"}), 404

        current_status = str(sheet.cell(row=row_id, column=11).value or "Checked In").strip()
        if current_status.lower() == "checked out":
            return jsonify({"status": "error", "message": "Visitor already checked out"}), 400

        timestamp = datetime.now()
        check_out_date = timestamp.strftime("%Y-%m-%d")
        check_out_time = timestamp.strftime("%H:%M:%S")

        sheet.cell(row=row_id, column=9).value = check_out_date
        sheet.cell(row=row_id, column=10).value = check_out_time
        sheet.cell(row=row_id, column=11).value = "Checked Out"
        workbook.save(visitor_workbook_filename)

        return jsonify({
            "status": "success",
            "message": f"Visitor checked out for {visitor_name}.",
            "row_id": row_id,
            "check_out_date": check_out_date,
            "check_out_time": check_out_time,
            "visitor_name": visitor_name,
        })
    except PermissionError:
        return jsonify({"status": "error", "message": "visitors.xlsx is open in Excel. Close it and try again."}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"Error: {str(e)}"}), 500
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass

def save_attendance_event(filename, record):
    try:
        initialize_excel(filename)
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook["Attendance"]
        result = upsert_attendance_event(sheet, record)
        workbook.save(filename)
        workbook.close()
        profile = get_employee_profile(name=record.get("name", ""), emp_id=record.get("emp_id", ""))
        sync_attendance_record_to_mongo({
            "name": record.get("name", ""),
            "emp_id": record.get("emp_id", ""),
            "department": (profile.get("department") or "General").strip() or "General",
            "date": record.get("date", ""),
            "time": record.get("time", ""),
            "check_in": result.get("check_in", ""),
            "check_out": result.get("check_out", ""),
            "total_hours": result.get("total_hours", ""),
            "status": result.get("status", ""),
            "in_status": result.get("in_status", ""),
            "out_status": result.get("out_status", ""),
            "notes": result.get("notes", ""),
            "confidence": record.get("confidence", ""),
        })
        return True, None, result
    except PermissionError:
        return False, "attendance.xlsx is open in Excel. Please close it and try again.", None
    except Exception as e:
        return False, str(e), None

if __name__ == '__main__':
    app.run(debug=True)

